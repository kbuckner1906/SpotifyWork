import json
import os

import requests

from exceptions import ResponseException
from secrets import spotify_token, spotify_user_id
from albums import Albums
import openpyxl
import math

 #Create A New Playlist
class Library:
    def __init__(self):
        self.trackids = []
        self.tracknames = []
        self.albumids = []
        self.ninetysalbums = []
        self.eightysalbums = []
        self.seventysalbums =[]
        self.canyougetanyolder = []
        self.thousands =[]
        self.wb = openpyxl.Workbook()
        self.genres = []
        self.artists = []
        self.anyyear=[]
        self.artistids = []
    def getTracks(self):
        # Initial Query
        query = "https://api.spotify.com/v1/me/tracks?limit=50&offset=0"
        response = requests.get(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # Initial Response
        json_response = response.json()
        items = json_response["items"]
        totalsongs = json_response["total"]
        print(totalsongs)
        remainder = totalsongs % 50
        runtimes = int(totalsongs/50)+1
        final = runtimes*50
        print(type(runtimes))
        count = 0
        i=0
        row=0
        songs = 0 
        #Loop
        while count < runtimes:
            if count == runtimes-1: 
                #while i < remainder:
                    trackname = items[i]['track']['name']
                    trackid = items[i]['track']['id']
                    trackuri= items[i]['track']['uri']
                    albumid = items[i]['track']['album']['id']
                    artistid = items[i]['track']['album']['artists'][0]["id"]
                    artist = items[i]['track']['album']['artists'][0]["name"]
                    print(trackname+ " by " + artist)
                    self.tracknames.append(trackname)
                    self.trackids.append(trackid)
                    self.albumids.append(albumid)
                    self.artists.append(artist)
                    self.artistids.append(artistid)
                    ws = self.wb.active
                    ws.cell(row=row+1, column=1).value = trackname
                    ws.cell(row=row+1, column=2).value = trackid
                    ws.cell(row=row+1, column=3).value = albumid
                    ws.cell(row=row+1, column=4).value = artistid
                    ws.cell(row=row+1, column=6).value = trackuri
                    ws.cell(row=row+1, column=8).value = artist
                    row = row +1
                    query = "https://api.spotify.com/v1/me/tracks?limit=" + str(remainder) + "&offset=" + str(final)
                    songs= songs+1
                    i = i + 1   
            else: 
                while i < 50:  
                    trackname = items[i]['track']['name']
                    trackid = items[i]['track']['id']
                    trackuri = items[i]['track']['uri']
                    albumid = items[i]['track']['album']['id']
                    artistid = items[i]['track']['album']['artists'][0]["id"]
                    artist = items[i]['track']['album']['artists'][0]["name"]
                    print(trackname+ " by " + artist)
                    self.tracknames.append(trackname)
                    self.trackids.append(trackid)
                    self.albumids.append(albumid)
                    self.artists.append(artistid)
                    self.artistids.append(artistid)
                    ws = self.wb.active
                    ws.cell(row=row+1, column=1).value = trackname
                    ws.cell(row=row+1, column=2).value = trackid
                    ws.cell(row=row+1, column=3).value = albumid
                    ws.cell(row=row+1,  column=8).value = artist
                    ws.cell(row=row+1,  column=6).value = trackuri
                    ws.cell(row=row+1, column=4).value = artistid
                    row = row +1
                    query = json_response['next']
                    songs = songs+1
                    i = i + 1
            count = count +1
        #New Query
            i=0
            response = requests.get(
                query,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer {}".format(spotify_token)
            })
            json_response = response.json()
            items = json_response["items"]
            print(totalsongs-songs)
            self.wb.save("Spotify Library.xlsx")
            print(str(count+1) + " out of " + str(runtimes+1) + " Done") 
    def create90s_playlist(self):
        request_body = json.dumps({
            "name": "That's So 90s",
            "description": "Songs from the 90s",
            "public": False
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]    
    def add90ssong_to_playlist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            if int(release_date) >= 1990 and int(release_date) <= 1999:
                albumid = ws.cell(row=i+1, column=6).value
                self.ninetysalbums.append(albumid)
                print("90's song added")
            i=i+1
        # collect all of uri
        uris = self.ninetysalbums
        # uri limit is 100 per request       
        request_len = math.ceil(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.create90s_playlist()
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))

            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)

            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_json = response.json()
            print("Successful Upload")
            print(response_json)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)
            else:
                end = end + 100
        return response_json
    def create80s_playlist(self):
        request_body = json.dumps({
            "name": "That's So 80s",
            "description": "Songs from the 80s",
            "public": False
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print("Playlist Created. You may now add your songs with the id below")
        print(response_json["id"])
        return response_json["id"]
    def add80ssong_to_playlist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            if int(release_date) >= 1980 and int(release_date) <= 1989:
                albumid = ws.cell(row=i+1, column=6).value
                self.eightysalbums.append(albumid)
                print("80's song added")
                print(i)
            i=i+1
        # collect all of uri
        uris = self.eightysalbums
        print(self.eightysalbums)
        # uri limit is 100 per request       
        request_len = round(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.create80s_playlist()
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 99
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)
            print(query)
            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            print(response.json())
            response_json = response.json()
            print("Successful Upload")
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)-1
            else:
                end = end + 100
        return response_json
    def create70s_playlist(self):
        request_body = json.dumps({
            "name": "That's So 70s",
            "description": "Songs from the 70s",
            "public": False
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]    
    def add70ssong_to_playlist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            if int(release_date) >= 1970 and int(release_date) <= 1979:
                albumid = ws.cell(row=i+1, column=6).value
                self.seventysalbums.append(albumid)
                print("70's song added")
            i=i+1
        # collect all of uri
        uris = self.seventysalbums
        # uri limit is 100 per request       
        request_len = math.ceil(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.create70s_playlist()
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))

            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)

            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_json = response.json()
            print("Successful Upload")
            print(response_json)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)
            else:
                end = end + 100
        return response_json
    def createoldies_playlist(self):
        request_body = json.dumps({
            "name": "That's So Ancient",
            "description": "Songs from 1969 and before",
            "public": False
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]      
    def addoldsong_to_playlist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            if int(release_date) <= 1969:
                albumid = ws.cell(row=i+1, column=6).value
                self.canyougetanyolder.append(albumid)
                print("Old Ass song added")
            i=i+1
        # collect all of uri
        uris = self.canyougetanyolder
        # uri limit is 100 per request       
        request_len = math.ceil(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.createoldies_playlist()
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))
            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)

            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
            print(query)
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_json = response.json()
            print("Successful Upload")
            print(response_json)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)
            else:
                end = end + 100
        return response_json
    def create00s_playlist(self):
        request_body = json.dumps({
            "name": "That's So 00s",
            "description": "Songs from the 00s",
            "public": False
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]    
    def add00ssongstoplaylist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            if int(release_date) >= 2000 and int(release_date) <= 2009:
                albumid = ws.cell(row=i+1, column=6).value
                self.thousands.append(albumid)
                print("Old Ass song added")
            i=i+1
        # collect all of uri
        uris = self.thousands
        # uri limit is 100 per request       
        request_len = math.ceil(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.create00s_playlist()
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))
            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)

            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_json = response.json()
            print("Successful Upload")
            print(response_json)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)
            else:
                end = end + 100
        return response_json     
    def libraryYears(self):
        wb = openpyxl.load_workbook("Spotify Library.xlsx")
        ws = wb.worksheets[0]
        i=0
       
        totalrows = ws.max_row
        lib = Albums()
        # count = int(totalrows/20)
        thisrow = 0 
        x=0
        while i < totalrows:
            yearsadded = 0
            while yearsadded < 50:
                if ws.cell(row=x+i+1, column=5).value is None: 
                    trackid= ws.cell(row=x+1, column=2).value
                    lib.albums.append(trackid)
                    x=x+1
                    yearsadded = yearsadded + 1
            print(lib.albums)
            res = list(filter(None, lib.albums))
            release_dates = lib.getreleasedates(res)
            for year in release_dates:
                ws.cell(row=thisrow+1, column=5).value = year
                thisrow+=1
            print(str(i+50)+ " Years have been added.")
            wb.save("Spotify Library Years.xlsx")
            lib.albums=[]
            i=i+50
        wb.save("Spotify Library Years.xlsx")       
    def getgenres(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        start = 0 
        end = 50
        genrecolumn = ws['G']
        testcolumn = ws['I']
        idcolumn = ws['D']  # Column
        column_list = [idcolumn[x].value for x in range(len(idcolumn))]
        namecolumn = ws['H']
        indexes = [column_list.index(x) for x in set(column_list)]
        uniqueartists = list(set(column_list)) 
        uniqueartistsnames = []
        for index in indexes:
            uniqueartistsnames.append(namecolumn[index].value)
        request_len = math.ceil(len(uniqueartists)/50)
        lib = Albums()
        uniqueartistgenres = []
        count = 0
        while i < request_len:
            search = uniqueartists[start:end]    
            genres = lib.getgenres(search)
            count = 0
            for genre in genres:
                #print(genre + " " + str(count+1))
                uniqueartistgenres.append(genre)
                count = count +1 
            print(len(uniqueartistgenres))
            count = 0
            start = start + 50
            if i == request_len-1:
               end = len(uniqueartists)
            else:
                end = end + 50
            print(str(end) + " Genres have been added.")
            wb.save("Spotify Library Years.xlsx")
            i=i+1
        for x in range(len(idcolumn)):
            index = uniqueartists.index(idcolumn[x].value)
            genrecolumn[x].value = uniqueartistgenres[index]
            testcolumn[x].value = uniqueartistsnames[index]
        wb.save("Spotify Library Years.xlsx")
    def createyearplaylist(self,year):
        request_body = json.dumps({
            "name": "That's So " + str(year),
            "description": "Songs from " + str(year),
            "public": False
        })
        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]    
    def addyearplaylist(self):
        year = int(input("What year would you like to choose? "))
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        i=0
        totalrows = ws.max_row
        while i < totalrows:
            release_date=ws.cell(row=i+1, column=5).value
            print(release_date)
            if int(release_date) == year:
                albumid = ws.cell(row=i+1, column=6).value
                self.anyyear.append(albumid)
            i=i+1
        # collect all of uri
        uris = self.anyyear
        print(uris)
        # uri limit is 100 per request       
        request_len = math.ceil(len(uris)/100) #4
        # create a new playlist
        playlist_id = self.createyearplaylist(year)
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = uris[begin:end]
            print(len(uris))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))
            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)
            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_jsons = response.json()
            print("Successful Upload")
            print(response_jsons)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(uris)
            else:
                end = end + 100
        #return response_jsons 
    def makeRBplaylist(self):
        wb = openpyxl.load_workbook("Spotify Library Years.xlsx")
        ws = wb.worksheets[0]
        genrecolumn = ws['G']
        uricolumn = ws['F']
        yearcolumn = ws['E']
        rnb = []
        genre_list = [genrecolumn[x].value for x in range(len(genrecolumn))]
        uri_list = [uricolumn[x].value for x in range(len(uricolumn))]
        year_list = [yearcolumn[x].value for x in range(len(yearcolumn))]
        count = 0
        for genre in genre_list:
            if "r&b" in genre and year_list[count] > 1990 and year_list[count] < 2000:
                rnb.append(uri_list[count])
            count= count+1
        request_len = math.ceil(len(rnb)/100)
        # create a new playlist
        playlist_id = self.createplaylist()
        print(len(rnb))
        print(playlist_id)
        # add all songs into new playlist
        i = 0
        begin = 0
        end = 100
        while i < request_len:
            thisuris = rnb[begin:end]
            print(len(rnb))
            print(len(thisuris))
            s = ","
            s = s.join(thisuris)
            print(len(s))
            query = "https://api.spotify.com/v1/playlists/{}/tracks?uris={}".format(
            playlist_id,s)
            print(query)
            response = requests.post(
            query,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        # check for valid response status
            if response.status_code != 201:
                raise ResponseException(response.status_code)
            response_jsons = response.json()
            print("Successful Upload")
            print(response_jsons)
            i = i+1
            begin = begin + 100
            if i == request_len-1:
               end = len(rnb)
            else:
                end = end + 100
            
    def createplaylist(self):
        request_body = json.dumps({
            "name": "Soul",
            "description": "Alternative RNB",
            "public": False
        })
        query = "https://api.spotify.com/v1/users/{}/playlists".format(
            spotify_user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_json = response.json()
        # playlist id
        print(response_json["id"])
        return response_json["id"]    
cp = Library()
#cp.getTracks()
#cp.libraryYears()
#cp.getgenres()
#cp.addoldsong_to_playlist()
#cp.add70ssong_to_playlist()
#cp.add80ssong_to_playlist()
#cp.add90ssong_to_playlist()
#cp.add00ssongstoplaylist()
#cp.addyearplaylist()
cp.makeRBplaylist()